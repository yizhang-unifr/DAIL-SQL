#!/usr/bin/env python3
"""End-to-end DAIL-SQL evaluation runner for the meteo domain.

Bypasses generate_question.py + ask_llm.py.  Builds prompts directly,
calls the project's LLMFactoryAdapter, optionally runs validator /
optimizer, evaluates via PostgreSQL, and saves results with XLSX export.

Usage (from project root):
    uv run src/DAIL-SQL/run/run_meteo.py \\
        --dataset src/DAIL-SQL/data/data_preprocess/test_data_point.json \\
        --ablation full \\
        --fewshot \\
        --end 10 \\
        --export-xlsx

Results are written to:
    src/DAIL-SQL/results/<model_short>/<ablation>/fewshot|no_fewshot/<YYYYMMDD_HHMMSS>/
        predictions.txt    one predicted SQL per line
        results.jsonl      per-question detail
        summary.json       EX score and counts
        report.xlsx        XLSX export (when --export-xlsx)
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

# ─── path setup ───────────────────────────────────────────────────────────────
_SCRIPT_DIR   = Path(__file__).resolve().parent          # src/DAIL-SQL/run/
_DAIL_ROOT    = _SCRIPT_DIR.parent                       # src/DAIL-SQL/
_PROJECT_ROOT = _DAIL_ROOT.parents[1]                    # ontology_retriever/
for _p in (str(_DAIL_ROOT), str(_PROJECT_ROOT)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from dotenv import load_dotenv
load_dotenv(_PROJECT_ROOT / ".env")

# ─── imports (all after path setup) ───────────────────────────────────────────
from meteo.geo_adapter import resolve_geo_from_places, format_geo_block
from meteo.meteo_entity_hint import build_entity_hint
from meteo.meteo_semantic_hint import build_semantic_hint
from meteo.llm_adapter import generate_candidates, get_adapter


def process_duplication(sql: str) -> str:
    return sql.strip().split("/*")[0]

# ─── meteo DB schema (static DDL for prompt) ──────────────────────────────────
_METEO_DDL = """\
CREATE TABLE meteo_tmean (
    time TIMESTAMP, latitude DOUBLE PRECISION, longitude DOUBLE PRECISION, tmean REAL
);
CREATE TABLE meteo_tmin (
    time TIMESTAMP, latitude DOUBLE PRECISION, longitude DOUBLE PRECISION, tmin REAL
);
CREATE TABLE meteo_tmax (
    time TIMESTAMP, latitude DOUBLE PRECISION, longitude DOUBLE PRECISION, tmax REAL
);
CREATE TABLE meteo_windspeedmax (
    time TIMESTAMP, latitude DOUBLE PRECISION, longitude DOUBLE PRECISION, windspeedmax REAL
);
CREATE TABLE meteo_tp (
    time TIMESTAMP, latitude DOUBLE PRECISION, longitude DOUBLE PRECISION, tp REAL
);
CREATE TABLE meteo_sdmax (
    time TIMESTAMP, latitude DOUBLE PRECISION, longitude DOUBLE PRECISION, sdmax REAL
);
CREATE TABLE meteo_elevation (
    latitude DOUBLE PRECISION, longitude DOUBLE PRECISION, elevation REAL
);
CREATE TABLE landcover_type (
    level3_code INTEGER, level3_label TEXT,
    level2_code INTEGER, level2_label TEXT,
    level1_code INTEGER, level1_label TEXT
);
CREATE TABLE landcover_upscaled (
    latitude DOUBLE PRECISION, longitude DOUBLE PRECISION,
    ranks INTEGER[][]
    -- ranks[i][1] = level3_code, ranks[i][2] = pixel_count
    -- iterate with: CROSS JOIN LATERAL GENERATE_SUBSCRIPTS(ranks, 1) AS i
);"""

_SCHEMA_BLOCK = f"/* Given the following database schema: */\n{_METEO_DDL}"

# ─── ablation config ──────────────────────────────────────────────────────────
_ABLATION_FLAGS = {
    "baseline": dict(geo=False, ogf=False, hints=False, validator=False, optimizer=False),
    "geo":      dict(geo=True,  ogf=False, hints=False, validator=False, optimizer=False),
    "ogf":      dict(geo=True,  ogf=True,  hints=False, validator=True,  optimizer=False),
    "hints":    dict(geo=True,  ogf=True,  hints=True,  validator=True,  optimizer=False),
    "full":     dict(geo=True,  ogf=True,  hints=True,  validator=True,  optimizer=True),
}

_COLUMN_CONTRACTS = {
    "meteo_tmean.tmean":               ["numeric", "unit:Kelvin — subtract 273.15 for Celsius"],
    "meteo_tmin.tmin":                 ["numeric", "unit:Kelvin — subtract 273.15 for Celsius"],
    "meteo_tmax.tmax":                 ["numeric", "unit:Kelvin — subtract 273.15 for Celsius"],
    "meteo_windspeedmax.windspeedmax": ["numeric", "unit:m/s"],
    "meteo_tp.tp":                     ["numeric", "unit:kg/m²/day"],
    "meteo_sdmax.sdmax":               ["numeric", "unit:metres"],
    "meteo_elevation.elevation":       ["numeric", "unit:metres above sea level"],
}


# ─── context loading ──────────────────────────────────────────────────────────

def _load_or_build_context(context_path: Path, dataset_path: Path) -> dict:
    """Load precomputed context, or build it on the fly if missing."""
    if context_path.exists():
        return json.loads(context_path.read_text()).get("by_question", {})

    print(f"[precompute] {context_path} not found — building now …")
    import subprocess
    cmd = [
        sys.executable, str(_DAIL_ROOT / "meteo" / "precompute_context.py"),
        "--questions", str(dataset_path),
        "--output", str(context_path),
    ]
    subprocess.run(cmd, check=True)
    return json.loads(context_path.read_text()).get("by_question", {})


def _load_fewshot(fewshot_path: Path) -> dict:
    if not fewshot_path.exists():
        print(f"[fewshot] {fewshot_path} not found — running without few-shot")
        return {}
    data = json.loads(fewshot_path.read_text(encoding="utf-8", errors="replace"))
    return data.get("by_question", {})


# ─── prompt builder ───────────────────────────────────────────────────────────

def _format_ogf_block(ogf_json: str) -> str:
    if not ogf_json:
        return ""
    try:
        ogf = json.loads(ogf_json)
    except Exception:
        return ""
    lines = ["/* Ontology-grounded function context:"]
    if tv := ogf.get("target_variable"):
        lines.append(f"   target_variable: {tv}")
    for spec in ogf.get("function_specs", []):
        if sem := spec.get("semantics", ""):
            lines.append(f"   semantics: {sem}")
        if tmpl := spec.get("pseudo_code_template", ""):
            lines.append(f"   formula: {tmpl}")
    lines.append("*/")
    return "\n".join(lines) + "\n"


def build_prompt(
    question: str,
    ctx: dict,
    flags: dict,
    fewshot_prompt: str,
) -> str:
    """Assemble the full DAIL-SQL style prompt."""
    blocks: list[str] = [_SCHEMA_BLOCK]

    if fewshot_prompt:
        blocks.append(fewshot_prompt)

    if flags["geo"]:
        pts = ctx.get("geo_points", [])
        if pts:
            blocks.append(format_geo_block(pts, mode="points").rstrip())

    if flags["ogf"]:
        ogf_block = _format_ogf_block(ctx.get("ogf_json", ""))
        if ogf_block:
            blocks.append(ogf_block.rstrip())

    if flags["hints"]:
        if eh := ctx.get("entity_hint", ""):
            blocks.append(eh.rstrip())
        if sh := ctx.get("semantic_hint", ""):
            blocks.append(sh.rstrip())

    blocks.append(f"/* Answer the following: {question} */\nSELECT ")
    return "\n\n".join(blocks)


# ─── SQL post-processing ──────────────────────────────────────────────────────

def _clean_sql(raw: str) -> str:
    """Strip thinking blocks, markdown fences, normalise to single-line SELECT."""
    # strip <think>...</think>
    raw = re.sub(r"<think>.*?</think>", "", raw, flags=re.DOTALL).strip()
    # strip markdown fences
    raw = re.sub(r"^```(?:sql)?|```$", "", raw, flags=re.MULTILINE).strip()
    # collapse extra whitespace
    sql = " ".join(raw.replace("\n", " ").split())
    sql = process_duplication(sql)
    upper = sql.upper().lstrip()
    if upper.startswith("SELECT") or upper.startswith("WITH"):
        return sql.lstrip()
    if sql.startswith(" "):
        return "SELECT" + sql
    return "SELECT " + sql


# ─── evaluation ───────────────────────────────────────────────────────────────

def _write_summary(
    run_dir: Path,
    args,
    model_short: str,
    dataset_path: Path,
    timestamp: str,
    total: int,
    correct: int,
    complete: bool,
) -> None:
    summary = {
        "ablation":        args.ablation,
        "model":           model_short,
        "fewshot":         args.fewshot,
        "dataset":         str(dataset_path),
        "n_questions":     total,
        "n_correct":       correct,
        "EX":              round(correct / total, 4) if total else 0.0,
        "timestamp":       timestamp,
        "complete":        complete,
    }
    (run_dir / "summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False)
    )


def _evaluate(predicted: str, gold: str, question_id: int | None = None) -> tuple[int, str]:
    """Return (1/0, error_msg). 1 = execution-equivalent."""
    try:
        from runner.database_manager import DatabaseManager
        resp = DatabaseManager.compare_sqls(
            predicted_sql=predicted,
            ground_truth_sql=gold,
            meta_time_out=60,
            question_id=question_id,
        )
        exec_res = 1 if resp.get("exec_res") == 1 else 0
        exec_err = resp.get("exec_err", "")
    except Exception as e:
        exec_res = 0
        exec_err = str(e)
    return exec_res, exec_err


# ─── XLSX export ──────────────────────────────────────────────────────────────

def _export_xlsx(records: list[dict], out_path: Path, ablation: str, model_name: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[xlsx] openpyxl not installed — skipping XLSX export")
        return

    wb = openpyxl.Workbook()

    # ── Results sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Results"
    headers = ["#", "question_id", "category", "question",
               "gold_sql", "predicted_sql", "exec_res", "exec_err"]
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=False)

    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")
    for i, rec in enumerate(records, 1):
        row = i + 1
        values = [
            i, rec.get("question_id", i - 1), rec.get("category", ""),
            rec["question"], rec["gold_sql"], rec["predicted_sql"],
            rec["exec_res"], rec.get("exec_err", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=False)
        # colour exec_res column
        res_cell = ws.cell(row=row, column=7)
        res_cell.fill = green if rec["exec_res"] == 1 else red

    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["H"].width = 40

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total = len(records)
    correct = sum(r["exec_res"] for r in records)
    ex = correct / total if total else 0.0
    for row, (k, v) in enumerate([
        ("Model",   model_name),
        ("Ablation", ablation),
        ("Total",    total),
        ("Correct",  correct),
        ("EX",       f"{ex:.1%}"),
    ], 1):
        ws2.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=v)

    wb.save(out_path)
    print(f"[xlsx] saved → {out_path}")


# ─── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="DAIL-SQL meteo evaluation runner")
    ap.add_argument("--dataset", default="",
                    help="Path to test JSON (default: src/DAIL-SQL/data/data_preprocess/test_data_point.json)")
    ap.add_argument("--ablation", choices=list(_ABLATION_FLAGS), default="full")
    ap.add_argument("--fewshot", action="store_true",
                    help="Inject few-shot examples from data/fewshot/questions.json")
    ap.add_argument("--end", type=int, default=10,
                    help="Max questions to evaluate (-1 = all)")
    ap.add_argument("--indices", type=str, default="",
                    help="Comma-separated question indices to evaluate (overrides --end)")
    ap.add_argument("--n", type=int, default=1,
                    help="Number of SQL candidates per question")
    ap.add_argument("--temperature", type=float, default=0.0)
    ap.add_argument("--meteo-context", type=str, default="",
                    help="Path to meteo_context.json (auto-built if absent)")
    ap.add_argument("--fewshot-path", type=str, default="",
                    help="Path to fewshot questions.json (default: src/DAIL-SQL/data/fewshot/questions.json)")
    ap.add_argument("--output-dir", type=str, default="",
                    help="Base results directory (default: src/DAIL-SQL/results)")
    ap.add_argument("--export-xlsx", action="store_true")
    ap.add_argument("--llm-config", type=str, default="",
                    help="Path to LLM config YAML (default: config/models.yaml)")
    args = ap.parse_args()

    if args.llm_config:
        from llm.model import set_llm_config_path
        set_llm_config_path(args.llm_config)

    dataset_path = Path(args.dataset).resolve() if args.dataset else (
        _DAIL_ROOT / "data" / "data_preprocess" / "test_data_point.json"
    )
    flags = _ABLATION_FLAGS[args.ablation]

    # ── resolve paths ──────────────────────────────────────────────────────────
    context_path = Path(args.meteo_context).resolve() if args.meteo_context else (
        _DAIL_ROOT / "data" / "meteo_context.json"
    )
    fewshot_path = Path(args.fewshot_path).resolve() if args.fewshot_path else (
        _DAIL_ROOT / "data" / "fewshot" / "questions.json"
    )
    base_results = Path(args.output_dir).resolve() if args.output_dir else (
        _DAIL_ROOT / "results"
    )

    # ── load data ──────────────────────────────────────────────────────────────
    all_items: list[dict] = json.loads(dataset_path.read_text())
    if args.indices:
        idx_set = [int(x.strip()) for x in args.indices.split(",")]
        items = [all_items[i] for i in idx_set]
    elif args.end != -1:
        items = all_items[: args.end]
    else:
        items = all_items
    print(f"Evaluating {len(items)} questions | ablation={args.ablation} | fewshot={args.fewshot}")

    # ── load gold SQL cache ────────────────────────────────────────────────────
    try:
        # Import via DAIL-SQL's execution module, which already handles the
        # OpenSearch-SQL sys.path injection internally.
        from runner.execution import set_gold_cache, _GoldSqlCache as GoldSqlCache
        if GoldSqlCache is None:
            raise ImportError("runner.gold_sql_cache not available (OpenSearch-SQL path not found)")
        _dataset_stem = dataset_path.stem  # e.g. "test_data_point"
        _cache_path = GoldSqlCache.cache_path(_PROJECT_ROOT, _dataset_stem, "meteo")
        _cache = GoldSqlCache(_cache_path)
        if _cache.is_loaded:
            set_gold_cache(_cache)
            print(f"Gold SQL cache loaded: {_cache_path} ({len(_cache._raw)} entries)")
        else:
            print(f"Gold SQL cache not found at {_cache_path}, falling back to live execution")
    except Exception as e:
        print(f"Warning: could not load gold SQL cache: {e}")

    # ── load context & fewshot ─────────────────────────────────────────────────
    ctx_dict: dict = {}
    if any(flags[k] for k in ("geo", "ogf", "hints", "validator")):
        ctx_dict = _load_or_build_context(context_path, dataset_path)

    fewshot_dict: dict = {}
    if args.fewshot:
        fewshot_dict = _load_fewshot(fewshot_path)

    # ── init LLM adapter + optional pipeline components ───────────────────────
    adapter = get_adapter()

    use_validator = flags["validator"]
    use_optimizer = flags["optimizer"]

    if use_validator:
        from pipeline.sql_audit.constraint_validator import validate_sql_candidates

    if use_optimizer:
        from pipeline.optimizer.detector import needs_optimization
        from pipeline.optimizer.mechanical import mechanical_optimize, verify_mechanical_transform
        from pipeline.optimizer.extract_rewriter import rewrite_extract_to_range, verify_extract_rewrite

    # ── get model name for output path ─────────────────────────────────────────
    try:
        from llm.model import _get_base_config
        model_short = _get_base_config().get("model", "unknown").split("/")[-1]
    except Exception:
        model_short = "unknown"

    # ── output directory ───────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fewshot_label = "fewshot" if args.fewshot else "no_fewshot"
    run_dir = base_results / model_short / args.ablation / fewshot_label / timestamp
    run_dir.mkdir(parents=True, exist_ok=True)
    print(f"Output → {run_dir}")

    # ── evaluation loop ────────────────────────────────────────────────────────
    records: list[dict] = []
    correct_so_far = 0

    llm_io_dir     = run_dir / "llm_io"
    llm_io_log_dir = llm_io_dir / "log"
    llm_io_log_dir.mkdir(parents=True, exist_ok=True)

    results_fh = open(run_dir / "results.jsonl", "w", encoding="utf-8")
    preds_fh   = open(run_dir / "predictions.txt", "w", encoding="utf-8")

    try:
        for i, item in enumerate(items):
            question = item["question"].strip()
            gold_sql  = item["SQL"].strip()
            q_key     = question.lower()
            category  = item.get("category", "")
            q_id      = item.get("question_id", i)

            print(f"\n[{i+1}/{len(items)}] {question[:70]}")

            # context lookup
            ctx = ctx_dict.get(q_key, {})

            # fewshot lookup
            fewshot_prompt = ""
            if args.fewshot:
                fewshot_prompt = fewshot_dict.get(q_key, {}).get("prompt", "")

            # build prompt
            prompt = build_prompt(question, ctx, flags, fewshot_prompt)

            # generate candidates
            t0 = time.time()
            candidates_raw = generate_candidates(prompt, n=args.n, temperature=args.temperature)
            elapsed = time.time() - t0
            candidates = [_clean_sql(c) for c in candidates_raw if c.strip()]
            if not candidates:
                candidates = ["SELECT NULL"]

            print(f"  LLM {elapsed:.1f}s | candidates={len(candidates)}")

            # validator
            if use_validator and candidates:
                ogf_payload: dict = {}
                if ctx.get("ogf_json"):
                    try:
                        ogf_payload = {"ontology_grounded_function": json.loads(ctx["ogf_json"])}
                    except Exception:
                        pass
                try:
                    candidates, _ = validate_sql_candidates(
                        adapter,
                        candidates,
                        _COLUMN_CONTRACTS,
                        ogf_payload,
                        question,
                    )
                    print(f"  validator ok")
                except Exception as e:
                    print(f"  validator error: {e}")

            # pick best candidate (first / only)
            final_sql = candidates[0] if candidates else "SELECT NULL"

            # mechanical optimizer (no LLM)
            sql_before_optimizer = final_sql
            optimizer_applied    = False
            optimizer_error      = ""
            if use_optimizer:
                try:
                    # Pass 1: VALUES subquery rewrite
                    if needs_optimization(final_sql):
                        rewritten, ok = mechanical_optimize(final_sql)
                        if ok and verify_mechanical_transform(final_sql, rewritten):
                            final_sql        = rewritten
                            optimizer_applied = True
                    # Pass 2: EXTRACT → time range rewrite
                    rewritten, ok = rewrite_extract_to_range(final_sql)
                    if ok and verify_extract_rewrite(final_sql, rewritten):
                        final_sql        = rewritten
                        optimizer_applied = True
                    if optimizer_applied:
                        print(f"  optimizer applied")
                except Exception as e:
                    optimizer_error = str(e)
                    print(f"  optimizer error: {e}")

            # evaluate
            exec_res, exec_err = _evaluate(final_sql, gold_sql, question_id=q_id)
            status = "✓" if exec_res else "✗"
            print(f"  {status} exec_res={exec_res}")

            if exec_res:
                correct_so_far += 1

            rec = {
                "question_id":   q_id,
                "category":      category,
                "question":      question,
                "gold_sql":      gold_sql,
                "predicted_sql": final_sql,
                "exec_res":      exec_res,
                "exec_err":      exec_err,
            }
            records.append(rec)

            # ── stream-write after every question ─────────────────────────────
            results_fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
            results_fh.flush()

            preds_fh.write(final_sql + "\n")
            preds_fh.flush()

            # ── per-query LLM I/O files ────────────────────────────────────────
            _SEP  = "=" * 70
            _LINE = "-" * 70
            geo_pts   = ctx.get("geo_points", [])
            geo_str   = json.dumps(geo_pts, ensure_ascii=False) if geo_pts else "(none)"
            raw_block = "\n---\n".join(candidates_raw) if candidates_raw else "(none)"
            opt_note  = (
                f"applied  before: {sql_before_optimizer}"
                if optimizer_applied else
                f"skipped{('  error: ' + optimizer_error) if optimizer_error else ''}"
            )
            file_stem = f"{i:04d}_{q_id}"

            # plain-text log — directly in llm_io/
            (llm_io_dir / f"{file_stem}.log").write_text(
                f"{_SEP}\n"
                f"Question [{i+1}/{len(items)}]  id={q_id}  category={category}\n"
                f"{question}\n"
                f"{_LINE}\n"
                f"[Context]\n"
                f"geo_points: {geo_str}\n"
                f"{ctx.get('entity_hint', '').strip()}\n"
                f"{ctx.get('semantic_hint', '').strip()}\n"
                f"ogf_json: {ctx.get('ogf_json', '(none)').strip()}\n"
                f"{_LINE}\n"
                f"[Human — Prompt]\n"
                f"{prompt}\n"
                f"{_LINE}\n"
                f"[AI — Raw Response]\n"
                f"{raw_block}\n"
                f"{_LINE}\n"
                f"[Optimizer]\n"
                f"{opt_note}\n"
                f"{_LINE}\n"
                f"[Final SQL]\n"
                f"{final_sql}\n"
                f"exec_res: {'✓' if exec_res else '✗'}  {exec_err}\n"
                f"{_SEP}\n",
                encoding="utf-8",
            )

            # structured JSON — in llm_io/log/
            (llm_io_log_dir / f"{file_stem}.json").write_text(
                json.dumps({
                    "question_id":        q_id,
                    "category":           category,
                    "question":           question,
                    "geo_points":         geo_pts,
                    "ogf_json":           ctx.get("ogf_json", ""),
                    "entity_hint":        ctx.get("entity_hint", ""),
                    "semantic_hint":      ctx.get("semantic_hint", ""),
                    "prompt":             prompt,
                    "raw_responses":      candidates_raw,
                    "sql_before_optimizer": sql_before_optimizer,
                    "optimizer_applied":  optimizer_applied,
                    "optimizer_error":    optimizer_error,
                    "final_sql":          final_sql,
                    "exec_res":           exec_res,
                    "exec_err":           exec_err,
                }, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )

            # partial summary (complete=False until loop ends)
            _write_summary(
                run_dir, args, model_short, dataset_path, timestamp,
                total=i + 1, correct=correct_so_far, complete=False,
            )

    finally:
        results_fh.close()
        preds_fh.close()

    # ── final summary ──────────────────────────────────────────────────────────
    total   = len(records)
    correct = correct_so_far
    ex      = correct / total if total else 0.0

    _write_summary(
        run_dir, args, model_short, dataset_path, timestamp,
        total=total, correct=correct, complete=True,
    )

    print(f"\n{'='*60}")
    print(f"EX = {correct}/{total} = {ex:.1%}  (ablation={args.ablation}, fewshot={args.fewshot})")
    print(f"Results → {run_dir}")

    # XLSX
    if args.export_xlsx:
        _export_xlsx(records, run_dir / "report.xlsx", args.ablation, model_short)


if __name__ == "__main__":
    main()
